"""Reads the uploaded spreadsheet into plain Python rows.

Parsing happens on the server, never in the browser, and streams row by row so a
large daily file does not have to be materialised twice.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.errors import ValidationError
from app.utils.text import clean

MAX_ROWS = 100_000


@dataclass(slots=True)
class ParsedSheet:
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)
    sheet_name: str | None = None
    skipped_blank_rows: int = 0

    def __len__(self) -> int:
        return len(self.rows)


def _dedupe_headers(raw_headers: list[Any]) -> list[str]:
    """Give every column a usable, unique name."""
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        name = clean(value) or f"Column {index + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 0
        headers.append(name)
    return headers


def parse_excel(
    payload: bytes, *, header_row: int = 1, sheet_name: str | None = None
) -> ParsedSheet:
    try:
        workbook = load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many shapes
        raise ValidationError(
            "The file could not be read as an Excel workbook. "
            "If it is an old .xls file, re-save it as .xlsx.",
            details=str(exc),
        ) from exc

    try:
        worksheet = (
            workbook[sheet_name]
            if sheet_name and sheet_name in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )

        headers: list[str] = []
        rows: list[dict[str, Any]] = []
        blank = 0

        for index, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index < header_row:
                continue
            if index == header_row:
                headers = _dedupe_headers(list(raw_row))
                continue
            if raw_row is None or all(cell is None or clean(cell) == "" for cell in raw_row):
                blank += 1
                continue
            record: dict[str, Any] = {}
            for position, header in enumerate(headers):
                record[header] = raw_row[position] if position < len(raw_row) else None
            record["__row__"] = index
            rows.append(record)
            if len(rows) >= MAX_ROWS:
                break

        if not headers:
            raise ValidationError(f"No header row was found at row {header_row} of the worksheet.")
        return ParsedSheet(
            headers=headers,
            rows=rows,
            sheet_name=worksheet.title,
            skipped_blank_rows=blank,
        )
    finally:
        workbook.close()


def parse_csv(payload: bytes, *, header_row: int = 1) -> ParsedSheet:
    text = _decode(payload)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    blank = 0

    for index, raw_row in enumerate(reader, start=1):
        if index < header_row:
            continue
        if index == header_row:
            headers = _dedupe_headers(raw_row)
            continue
        if not raw_row or all(clean(cell) == "" for cell in raw_row):
            blank += 1
            continue
        record: dict[str, Any] = {}
        for position, header in enumerate(headers):
            record[header] = raw_row[position] if position < len(raw_row) else None
        record["__row__"] = index
        rows.append(record)
        if len(rows) >= MAX_ROWS:
            break

    if not headers:
        raise ValidationError(f"No header row was found at line {header_row}.")
    return ParsedSheet(headers=headers, rows=rows, skipped_blank_rows=blank)


def _decode(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def parse(
    payload: bytes,
    filename: str,
    *,
    header_row: int = 1,
    sheet_name: str | None = None,
) -> ParsedSheet:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return parse_csv(payload, header_row=header_row)
    if extension == ".xls":
        raise ValidationError(
            "Legacy .xls files are not supported directly. "
            "Please open the file in Excel and save it as .xlsx or .csv."
        )
    return parse_excel(payload, header_row=header_row, sheet_name=sheet_name)


def iter_rows(sheet: ParsedSheet) -> Iterator[tuple[int, dict[str, Any]]]:
    for record in sheet.rows:
        row_number = int(record.get("__row__", 0))
        data = {k: v for k, v in record.items() if k != "__row__"}
        yield row_number, data
