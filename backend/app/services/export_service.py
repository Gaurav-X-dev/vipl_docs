"""XLSX / CSV export.

Exports always honour the filters visible on screen — the caller passes the same
:class:`CaseFilters` the list endpoint used, so what you see is what you get.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Sequence
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.dates import format_date, format_datetime, utcnow

HEADER_FILL = PatternFill("solid", fgColor="FFC000")
HEADER_FONT = Font(bold=True, color="1F2937")


def _stringify(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return format_datetime(value) if hasattr(value, "hour") else format_date(value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def to_xlsx(
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    sheet_title: str = "Export",
    freeze_header: bool = True,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]

    sheet.append(list(headers))
    for column_index in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    widths = [len(str(h)) + 2 for h in headers]
    for row in rows:
        values = [_stringify(value) for value in row]
        sheet.append(values)
        for index, value in enumerate(values):
            widths[index] = min(60, max(widths[index], len(str(value)) + 2))

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    if freeze_header:
        sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_csv(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(list(headers))
    for row in rows:
        writer.writerow([_stringify(value) for value in row])
    # BOM so Excel opens UTF-8 correctly on Windows.
    return buffer.getvalue().encode("utf-8-sig")


def filename(prefix: str, extension: str, when: date | None = None) -> str:
    stamp = (when or utcnow().date()).strftime("%Y-%m-%d")
    return f"{prefix}_{stamp}.{extension}"


#: Column order mirrors the client's own daily file (Image 3) so an export can be
#: fed straight back into their process.
CASE_EXPORT_HEADERS: tuple[str, ...] = (
    "Case No",
    "Co. Name",
    "Case Type",
    "Category",
    "Month",
    "Date",
    "Aging",
    "KRN No",
    "Policy Number",
    "Application Number",
    "Life_Assured_Name",
    "City",
    "State",
    "Pin Code",
    "Assign To",
    "Status",
    "Outcome",
    "Report Status",
    "Priority",
    "TAT State",
    "Due Date",
    "Report Date",
    "Completion Date",
    "Report Prep By",
    "Remark/ADD IO ID",
)


def case_export_row(case: Any, aging: int | None, tat_state: str) -> list[Any]:
    return [
        case.case_number,
        case.company.short_name if case.company else "",
        case.case_type.name if case.case_type else "",
        case.category.value.replace("_", " ").title(),
        case.received_month or format_date(case.received_at, "%b-%Y"),
        format_date(case.received_at),
        aging if aging is not None else "",
        case.krn_no or "",
        case.policy_number or "",
        case.application_number or "",
        case.life_assured_name,
        case.city or "",
        case.state or "",
        case.pin_code or "",
        case.assigned_to.full_name if case.assigned_to else "",
        case.status.value.replace("_", " ").title(),
        case.outcome.value.title() if case.outcome else "",
        case.report_status.value.title() if case.report_status else "",
        case.priority.value.title(),
        tat_state.replace("_", " ").title(),
        format_date(case.due_at),
        format_date(case.report_date),
        format_date(case.completion_date or case.completed_at),
        case.report_prepared_by or "",
        case.import_remark or "",
    ]


IMPORT_ERROR_HEADERS: tuple[str, ...] = (
    "Row No",
    "Status",
    "Error Message",
    "Warnings",
)


def import_error_workbook(rows: Iterable[Any], original_headers: list[str]) -> bytes:
    """Downloadable rejected-rows workbook, with the original data preserved."""
    headers = list(IMPORT_ERROR_HEADERS) + original_headers
    body: list[list[Any]] = []
    for row in rows:
        raw = row.raw_data or {}
        body.append(
            [
                row.row_number,
                row.status.value.title(),
                "; ".join(row.errors or []),
                "; ".join(row.warnings or []),
                *[raw.get(header, "") for header in original_headers],
            ]
        )
    return to_xlsx(headers, body, sheet_title="Rejected rows")
